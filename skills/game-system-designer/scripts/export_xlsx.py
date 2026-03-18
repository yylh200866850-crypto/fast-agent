#!/usr/bin/env python3
"""
游戏UI原型图分析 → Excel需求清单导出

用法: python export_xlsx.py --data analysis.json [--image prototype.png] --output output.xlsx

生成包含以下Sheet的Excel文件:
  1. 组件清单 - 所有UI组件标注表
  2. 前端需求 - Client Side需求
  3. 服务器需求 - Server Side需求
  4. 美术需求 - Art Side需求
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[错误] 请先安装 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl.drawing.image import Image as XlImage
    HAS_IMAGE = True
except ImportError:
    HAS_IMAGE = False

# ============ 样式 ============
HEADER_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
DATA_FONT = Font(name='Microsoft YaHei', size=10)
LABEL_FONT = Font(name='Microsoft YaHei', bold=True, size=10, color='1F4E79')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
ALT_FILL = PatternFill('solid', fgColor='F2F7FB')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
WRAP = Alignment(vertical='center', wrap_text=True)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data(ws, row, cols, is_alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = ALT_FILL


def auto_width(ws, min_width=12, max_width=50):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        lengths = [len(str(cell.value or '')) for cell in col]
        width = min(max(max(lengths) + 4, min_width), max_width)
        ws.column_dimensions[letter].width = width


# ============ Sheet: 组件清单 ============
def build_components_sheet(wb, data, image_path=None):
    ws = wb.active
    ws.title = "组件清单"

    # 插入原型图
    start_row = 1
    if image_path and HAS_IMAGE and Path(image_path).exists():
        try:
            img = XlImage(image_path)
            img.width = 600
            img.height = 400
            ws.add_image(img, 'A1')
            start_row = 24  # 留出图片空间
        except Exception:
            pass

    headers = ['编号', '组件名称', '组件类型', '位置描述', '尺寸估算', '功能说明']
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header(ws, start_row, len(headers))

    for i, comp in enumerate(data.get('components', []), 1):
        r = start_row + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=comp.get('name', ''))
        ws.cell(row=r, column=3, value=comp.get('type', ''))
        ws.cell(row=r, column=4, value=comp.get('position', ''))
        ws.cell(row=r, column=5, value=comp.get('size', ''))
        ws.cell(row=r, column=6, value=comp.get('description', ''))
        style_data(ws, r, len(headers), is_alt=(i % 2 == 0))
        ws.cell(row=r, column=1).alignment = CENTER

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 40


# ============ Sheet: 需求列表 ============
def build_requirement_sheet(wb, sheet_name, section_data):
    ws = wb.create_sheet(sheet_name)
    headers = ['类别', '序号', '需求描述', '优先级', '备注']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for category, items in section_data.items():
        if not isinstance(items, list):
            items = [items]
        for idx, item in enumerate(items, 1):
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=idx)
            ws.cell(row=row, column=3, value=item)
            ws.cell(row=row, column=4, value='P1')  # 默认优先级
            ws.cell(row=row, column=5, value='')
            style_data(ws, row, len(headers), is_alt=(row % 2 == 0))
            ws.cell(row=row, column=1).font = LABEL_FONT
            row += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 20


# ============ 主流程 ============
def main():
    parser = argparse.ArgumentParser(description="导出游戏UI分析Excel文档")
    parser.add_argument("--data", required=True, help="分析数据JSON文件")
    parser.add_argument("--image", help="原型图路径")
    parser.add_argument("--output", "-o", default="output.xlsx", help="输出Excel路径")
    args = parser.parse_args()

    data = json.loads(Path(args.data).read_text(encoding='utf-8'))

    wb = Workbook()
    build_components_sheet(wb, data, args.image)

    if data.get('clientSide'):
        build_requirement_sheet(wb, '前端需求', data['clientSide'])
    if data.get('serverSide'):
        build_requirement_sheet(wb, '服务器需求', data['serverSide'])
    if data.get('artSide'):
        build_requirement_sheet(wb, '美术需求', data['artSide'])

    wb.save(args.output)
    print(f"[OK] Excel文档已生成: {args.output}")


if __name__ == "__main__":
    main()
